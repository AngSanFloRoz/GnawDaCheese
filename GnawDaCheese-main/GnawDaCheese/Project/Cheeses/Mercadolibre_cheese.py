from time import sleep # para que se quede esperando en sleep(t) t son segundos
import json

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys # send keys in navegator

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

#class onlinecommerce only fot MERCADO LIBRE

import os

class Onlinecommerce:
    def __init__(self, url: str):
        base_dir = os.path.dirname(os.path.abspath(__file__))  # ruta del archivo actual
        chromedriver_path = os.path.join(base_dir, "..", "chromedriver-linux64", "chromedriver")
        self.service = Service(executable_path=chromedriver_path)
        self.driver = webdriver.Chrome(service=self.service)
        self.url = url

    def gnaw(self):
        ''' with driver open navegator'''
        return self.driver.get(self.url)
    

    def buscar(self, palabra: str):
        self.gnaw()
        self.palabra = palabra
        barra = self.driver.find_element(By.NAME, "as_word")
        barra.send_keys(palabra)
        sleep(1)
        barra.send_keys(Keys.RETURN)

    def gnaw_products(self, palabra):
        self.buscar(palabra)
        sleep(5)
        casillas = self.driver.find_elements(
            By.CLASS_NAME, "ui-search-layout__item")
        datost = []
        sleep(5)
        for producto in casillas:  #only 10
            #* title component
            
            name = producto.find_element(
                By.CLASS_NAME, "poly-component__title").text
            link = producto.find_element(
                By.TAG_NAME, "a").get_attribute("href")

            #* get brand and seller

            try:
                brand = producto.find_element(
                    By.CLASS_NAME, "poly-component__brand").text
            except Exception:
                brand = "Not found"
            try:
                seller = producto.find_element(
                    By.CLASS_NAME, "poly-component__seller").text
            except Exception:
                seller = "Not found"
            
            #* reviews component

            try:
                reviews = producto.find_element(
                    By.CLASS_NAME, "poly-component__reviews")
                
                rating = reviews.find_element(
                    By.CLASS_NAME, "poly-reviews__rating").text
                
                total_reviews =  reviews.find_element(
                    By.CLASS_NAME, "poly-reviews__total").text
            
            except Exception:
                rating = "No reviews found"
                total_reviews = rating


            #* price component 
            price = producto.find_element(
                By.CLASS_NAME, "poly-component__price")
            
            try:    
                pricein = price.find_element(
                    By.CSS_SELECTOR, ".andes-money-amount.andes-money-amount--previous.andes-money-amount--cents-comma")
                
                priceaft = pricein.find_element(
                    By.CLASS_NAME, "andes-money-amount__fraction").text
            
            except Exception:
                priceaft = None 

            mainprice = price.find_element(
                By.CLASS_NAME, "poly-price__current")
            
            price = mainprice.find_element(
                By.CLASS_NAME, "andes-money-amount__fraction").text
                
            if priceaft:
                discount = mainprice.find_element(
                    By.CLASS_NAME, "andes-money-amount__discount").text
                datost.append((name, brand, seller, priceaft, price, discount
                               , rating, total_reviews, link)) 
            else:
                datost.append((name, brand, seller, price, price, "No discount"
                               ,rating, total_reviews, link))    
        # tried to print
        #    for Title, brand, seller, precioaft, price, discount, rating, total_reviews, link in datost:
        #        print(f"Product: {Title}| brand: {brand} / seller: {seller} | SubPrice: ${precioaft} | price: {price}| discount: {discount} 
        #        | rating: {rating} | total_reviews: {total_reviews} | {link}")
        #    sleep(60)
        return datost
    
    def to_dict(self, listproducts: list):
        '''Convert scrapped data to dict for convert it in a json file'''
        listdict = []
        for p in listproducts:
            name, brand, seller, priceaft, price, discount, rating, total_reviews, link = p
            diccionario = {
                "Name": name,
                "Brand": brand,
                "Seller": seller,
                "Subprice": priceaft,
                "Price": price,
                "Discount": discount,
                "Rating": rating,
                "Total reviews": total_reviews,
                "Link": link,
            }
            listdict.append(diccionario)
        return listdict
    
    def save_to_json(self, filename, data):
        '''Guardar datos en formato JSON en Project/jsonfilesdinamic'''
        self.json_name = filename
        # Ruta del archivo actual
        base_dir = os.path.dirname(__file__)

        # Subes un nivel a Project y entras a jsonfilesdinamic
        json_dir = os.path.join(base_dir, "..", "jsonfilesdinamic")
        os.makedirs(json_dir, exist_ok=True)  # Crea la carpeta si no existe

        # Ruta final del archivo
        json_path = os.path.join(json_dir, f"{filename}.json")

        # Guardar el JSON
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        
    def json_a_excel(self, nombre_excel):
        # Palabras clave que deben ir en negrilla
        palabras_negrita = ["No discount", "No reviews found", "Not found"]  

        # Obtener ruta absoluta al archivo actual
        base_dir = os.path.dirname(__file__)

        # Ruta al JSON (subes a Project/jsonfilesdinamic)
        json_path = os.path.join(base_dir, "..", "jsonfilesdinamic", f"{self.json_name}.json")
    
        with open(json_path, "r", encoding="utf-8") as f:
            datos = json.load(f)

        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = self.palabra

        # Agregar encabezados
        headers = list(datos[0].keys())
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Agregar datos con negrilla condicional
        for item in datos:
            row = []
            for col in headers:
                valor = str(item[col])
                fuente = Font(bold=any(palabra.lower() in valor.lower() for palabra in palabras_negrita))
                row.append((valor, fuente))

            # Escribir valores
            ws.append([v[0] for v in row])
            for idx, cell_info in enumerate(row, start=1):
                celda = ws.cell(row=ws.max_row, column=idx)
                celda.font = cell_info[1]

        # Ajustar ancho de columnas y alinear contenido
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[col_letter].width = max_length + 2

        # Congelar encabezado
        ws.freeze_panes = "A2"

        # Crear carpeta de destino si no existe
        excel_dir = os.path.join(base_dir, "..", "excelfilesdinamic")
        os.makedirs(excel_dir, exist_ok=True)

        # Guardar Excel
        excel_path = os.path.join(excel_dir, f"{nombre_excel}.xlsx")
        wb.save(excel_path)
        print(f"Archivo Excel guardado en: {excel_path}")

    def scrapped(self, search, filename):
        a = webd1.gnaw_products(search)
        b = webd1.to_dict(a)
        webd1.save_to_json(filename, b)
        webd1.json_a_excel(filename)
    
        

url = "https://www.mercadolibre.com.co/"
webd1 = Onlinecommerce(url)
webd1.scrapped("sans", "e_e_e")
